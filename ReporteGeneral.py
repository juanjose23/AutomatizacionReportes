import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

#Rutas de los datos  de Archivo
ArchivoCompra="Data/Pedido de compra (purchase.order) (1).xlsx"
ArchivoVenta="Data/Pedido de venta(purchase.order) (1) .xlsx"
ArchivoPrecio="Data/Precios.xlsx"

#Obtener la fecha de hoy
fecha_hoy = datetime.now().strftime('%Y-%m-%d')

# Definir el nombre del archivo con la fecha de hoy
output_file = f'Excel/Reporte_{fecha_hoy}.xlsx'

# Importar los datos de ArchivoCompra
df_compra = pd.read_excel(ArchivoCompra)
resumenCompra = df_compra.groupby('Líneas del pedido/Nombre mostrado')['Líneas del pedido/Cantidad total'].sum().reset_index()
resumenCompra.rename(columns={'Líneas del pedido/Cantidad total': 'INVENTARIO TEÓRICO'}, inplace=True)

# Importar los datos de ArchivoVenta
df_venta = pd.read_excel(ArchivoVenta)
resumenVenta = df_venta.groupby('Líneas del pedido/Nombre mostrado')['Líneas del pedido/Cantidad total'].sum().reset_index()
resumenVenta.rename(columns={'Líneas del pedido/Cantidad total': 'INVENTARIO FÍSICO'}, inplace=True)

# Combinar los resultados en una tabla final
inventario_comparado = pd.merge(resumenCompra, resumenVenta, left_on='Líneas del pedido/Nombre mostrado', right_on='Líneas del pedido/Nombre mostrado', how='outer')

# Renombrar la columna para que coincida con el formato deseado
inventario_comparado.rename(columns={'Líneas del pedido/Nombre mostrado': 'PRODUCTO'}, inplace=True)

# Reordenar las columnas según el formato deseado
inventario_comparado = inventario_comparado[['PRODUCTO', 'INVENTARIO FÍSICO', 'INVENTARIO TEÓRICO']]

# Rellenar valores NaN con 0 si es necesario
inventario_comparado.fillna(0, inplace=True)

# Calcular la diferencia y agregarla como una nueva columna
inventario_comparado['DIFERENCIA'] = inventario_comparado['INVENTARIO FÍSICO'] - inventario_comparado['INVENTARIO TEÓRICO']

# Importar los datos de ArchivoPrecios
df_precios = pd.read_excel(ArchivoPrecio, sheet_name=0)
df_precios = df_precios[['Elementos de tarifa/Plantilla de producto/Nombre', 'Elementos de tarifa/Precio fijo']]
df_precios.rename(columns={'Elementos de tarifa/Plantilla de producto/Nombre': 'PRODUCTO', 'Elementos de tarifa/Precio fijo': 'PRECIO'}, inplace=True)

# Importar la segunda hoja con la clasificación de material
df_materiales = pd.read_excel(ArchivoPrecio, sheet_name=1)
df_materiales = df_materiales[['PRODUCTO', 'MATERIAL']]

# Combinar inventario_comparado con df_precios
inventario_comparado = pd.merge(inventario_comparado, df_precios, on='PRODUCTO', how='left')

# Agregar la columna MATERIAL
inventario_comparado = pd.merge(inventario_comparado, df_materiales, on='PRODUCTO', how='left')

# Calcular la columna MONETARIO
inventario_comparado['MONETARIO'] = inventario_comparado['DIFERENCIA'] * inventario_comparado['PRECIO']

# Seleccionar las columnas a mostrar
inventario_comparado = inventario_comparado[['PRODUCTO', 'INVENTARIO FÍSICO', 'INVENTARIO TEÓRICO', 'DIFERENCIA', 'MONETARIO', 'MATERIAL']]

# Guardar el resultado en un archivo Excel con estilo
inventario_comparado.to_excel(output_file, sheet_name='POWER BI', index=False)
# Estilo de la tabla corporativa
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Color de encabezado
even_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Color de filas pares
odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Color de filas impares

border_side = Side(border_style="thin", color="000000")  # Borde fino negro
border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

font_header = Font(bold=True, color="FFFFFF")  # Fuente blanca en negrita para encabezados

# Crear una tabla en la hoja de Excel
wb = load_workbook(output_file)
ws = wb['POWER BI']

table = Table(displayName='InventarioComparado', ref=ws.dimensions)

# Asignar el estilo a la tabla
style = TableStyleInfo(
    name='TableStyleLight2', showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=True
)
table.tableStyleInfo = style

# Aplicar formato de fondo y bordes a las filas y columnas
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.row % 2 == 0:
            cell.fill = even_row_fill
        else:
            cell.fill = odd_row_fill
        cell.border = border

# Estilo de los encabezados
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = font_header
    cell.border = border

# Añadir la tabla a la hoja
ws.add_table(table)

# Guardar el archivo con la tabla y filtros
wb.save(output_file)

print(f"Archivo guardado como '{output_file}' con la pestaña 'POWER BI' y tabla con filtros.")